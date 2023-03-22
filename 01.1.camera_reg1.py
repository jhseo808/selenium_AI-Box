from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

import time
import openpyxl
import os
try:
    os.makedirs('D:\Auto_test')
    print('Auto테스트 파일이 생성되었습니다.')
except FileExistsError as a:
    print(a)

# sys.stdout = open('D:\Auto_test\AI-Box_log.txt','w') # print Log


#브라우저 꺼짐 방지
options = Options()
options.add_experimental_option("detach", True)
service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)
driver.implicitly_wait(3)
driver.maximize_window()

# 엑셀 시트 경로
fpath = r'C:\startcoding\AIBox.xlsx'
wb = openpyxl.load_workbook(fpath)
# ws = wb.active
ws = wb['AI-Box_Info']


driver.get('http://172.16.6.230/web/#/web/auth')
action = ActionChains(driver)

driver.find_element(By.CLASS_NAME, 'mat-form-field-infix')
(action.send_keys(ws['A2'].value).key_down(Keys.TAB).send_keys(ws['B2'].value).pause(1).key_down(Keys.ENTER).perform())
action.reset_actions
time.sleep(1)
print('Login OK\n----------------------------------------------------------------')

driver.get('http://172.16.6.230/web/#/web/camera-setting') # 카메라 등록 페이지

# 카메라 삭제
driver.find_element(By.XPATH, \
     '/html/body/app-root/app-sidenav-responsive/div/mat-sidenav-container/mat-sidenav-content/app-camera-setting/div/mat-card/form/mat-table/mat-row[1]/mat-cell[1]/button'\
         ).click()
print('Ch1 delete')
driver.find_element(By.XPATH, \
     '/html/body/app-root/app-sidenav-responsive/div/mat-sidenav-container/mat-sidenav-content/app-camera-setting/div/mat-card/form/mat-table/mat-row[3]/mat-cell[1]/button'\
         ).click()
print('Ch2 delete')
driver.find_element(By.XPATH, \
    '/html/body/app-root/app-sidenav-responsive/div/mat-sidenav-container/mat-sidenav-content/app-camera-setting/div/mat-card/form/mat-table/mat-row[5]/mat-cell[1]/button'\
       ).click()
print('Ch3 delete')
driver.find_element(By.XPATH, \
   '/html/body/app-root/app-sidenav-responsive/div/mat-sidenav-container/mat-sidenav-content/app-camera-setting/div/mat-card/form/mat-table/mat-row[7]/mat-cell[1]/button'\
       ).click()
print('Ch4 delete')

driver.find_element(By.CSS_SELECTOR, \
   'body > app-root > app-sidenav-responsive > div > mat-sidenav-container > mat-sidenav-content > app-camera-setting > div > mat-card > form > div > app-btn-apply > button'\
       ).click()
print('All Ch delete complete!\n------------------------------------------------------------')

ch1 = driver.find_element(By.CSS_SELECTOR, '#mat-input-12')
ch1.click()
(
    action.send_keys('Test!테스트@#123').key_down(Keys.TAB).send_keys(ws['D3'].value).key_down(Keys.TAB)
    .send_keys(ws['A3'].value).key_down(Keys.TAB).send_keys(ws['B3'].value).perform()
    )
print('Ch1 register')

ch2 = driver.find_element(By.CSS_SELECTOR, '#mat-input-18')
ch2.click()
(
    action.send_keys('QA!한글00000456').key_down(Keys.TAB).send_keys(ws['D4'].value).key_down(Keys.TAB)
    .send_keys(ws['A4'].value).key_down(Keys.TAB).send_keys(ws['B4'].value).perform()
    )
print('Ch2 register')
ch3 = driver.find_element(By.CSS_SELECTOR, '#mat-input-24')
ch3.click()
(
    action.send_keys('연결999').key_down(Keys.TAB).send_keys(ws['D5'].value).key_down(Keys.TAB)
    .send_keys(ws['A5'].value).key_down(Keys.TAB).send_keys(ws['B5'].value).perform()
    )
print('Ch3 register')
ch4 = driver.find_element(By.CSS_SELECTOR, '#mat-input-30')
ch4.click()
(
    action.send_keys('test999').key_down(Keys.TAB).send_keys(ws['D6'].value).key_down(Keys.TAB)
    .send_keys(ws['A6'].value).key_down(Keys.TAB).send_keys(ws['B6'].value).perform()
    )
ch4PTZlist = driver.find_element(By.CSS_SELECTOR, 'body > app-root > app-sidenav-responsive > div > mat-sidenav-container > mat-sidenav-content > app-camera-setting > div > mat-card > form > mat-table > mat-row:nth-child(9) > mat-cell.mat-cell.cdk-column-conType.mat-column-conType.ng-star-inserted > mat-form-field')
ch4PTZlist.click()
ch4PTZ = driver.find_element(By.CSS_SELECTOR, '#mat-option-157 > span')
ch4PTZ.click()
print('Ch4 PTZ register') 

get_cells = ws['A3':'D6']
for row in get_cells:
    for cell in row:
        print(cell.value, end=' ')
    print() # 카메라별 등록 정보
